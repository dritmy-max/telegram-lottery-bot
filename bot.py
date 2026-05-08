import os
import logging
import openpyxl
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes

# Настройка логирования (чтобы видеть, что происходит)
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ВАЖНО: Вставьте ваш токен
TOKEN = "8718532267:AAEq7afHk_Nuqjy3KeqI52KdzanQLQ1_iEI"
EXCEL_FILE = "lottery.xlsx"

# --- Работа с Excel ---
def init_excel():
    """Создает файл Excel, если его нет"""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        # Лист с номерами
        sheet = wb.active
        sheet.title = "Номера"
        sheet.append(["Номер", "Свободен", "Владелец (ID)"])
        for i in range(1, 401):
            sheet.append([i, "✅", ""])
        # Лист с участниками
        ws2 = wb.create_sheet("Участники")
        ws2.append(["Telegram ID", "Телефон", "Купленные номера"])
        wb.save(EXCEL_FILE)
        logger.info("Excel файл создан: %s", EXCEL_FILE)
    else:
        logger.info("Excel файл уже существует")

def get_available_numbers():
    """Возвращает список свободных номерков"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb["Номера"]
        available = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == "✅":
                available.append(row[0])
        wb.close()
        return available
    except Exception as e:
        logger.error("Ошибка при чтении Excel: %s", e)
        return []

def buy_numbers(user_id, phone, numbers):
    """Покупает номера, записывает в Excel"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb["Номера"]
        # Помечаем купленные
        for row in sheet.iter_rows(min_row=2, max_row=401, values_only=False):
            if row[0].value in numbers and row[1].value == "✅":
                row[1].value = "❌"
                row[2].value = str(user_id)
        # Записываем участника
        ws2 = wb["Участники"]
        ws2.append([str(user_id), phone, ", ".join([str(n) for n in numbers])])
        wb.save(EXCEL_FILE)
        logger.info("Куплены номера: %s для user %s", numbers, user_id)
        return True
    except Exception as e:
        logger.error("Ошибка при покупке: %s", e)
        return False

# --- Команды бота ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🎫 Добро пожаловать в лотерею!\n"
        "Команды:\n"
        "/buy - купить номера\n"
        "/available - показать свободные номера\n"
        "/my_tickets - мои номера\n"
        "/draw - розыгрыш (только для админа)"
    )

async def available(update: Update, context: ContextTypes.DEFAULT_TYPE):
    free = get_available_numbers()
    if not free:
        await update.message.reply_text("🎟 Все номера распроданы!")
        return
    text = "🎫 Свободные номера:\n"
    for i in range(0, len(free), 20):
        text += ", ".join([str(n) for n in free[i:i+20]]) + "\n"
    await update.message.reply_text(text)

async def buy_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    free = get_available_numbers()
    if not free:
        await update.message.reply_text("🤷 Все номера проданы!")
        return
    keyboard = []
    row = []
    for num in free[:50]:
        row.append(InlineKeyboardButton(str(num), callback_data=f"buy_{num}"))
        if len(row) == 10:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите номер для покупки:", reply_markup=reply_markup)

async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data.startswith("buy_"):
        num = int(data.split("_")[1])
        user_id = query.from_user.id
        # В реальном боте здесь надо запросить телефон
        success = buy_numbers(user_id, "не указан", [num])
        if success:
            await query.edit_message_text(f"✅ Номер {num} куплен!\nСпасибо за участие!")
        else:
            await query.edit_message_text("❌ Ошибка при покупке. Попробуйте позже.")

# --- Запуск бота ---
def main():
    logger.info("Запуск бота...")
    init_excel()
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("available", available))
    app.add_handler(CommandHandler("buy", buy_handler))
    app.add_handler(CallbackQueryHandler(button_callback))
    logger.info("Бот запущен и готов к работе")
    app.run_polling()

if __name__ == "__main__":
    main()
